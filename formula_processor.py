import openpyxl
import re
from typing import List, Dict, Any
import traceback

class FormulaProcessor:
    """
    Processes Excel workbooks to extract and analyze formulas
    """
    
    def __init__(self):
        self.formula_patterns = {
            'VLOOKUP': r'VLOOKUP\s*\(',
            'HLOOKUP': r'HLOOKUP\s*\(',
            'INDEX': r'INDEX\s*\(',
            'MATCH': r'MATCH\s*\(',
            'IF': r'IF\s*\(',
            'SUMIF': r'SUMIF\s*\(',
            'COUNTIF': r'COUNTIF\s*\(',
            'CONCATENATE': r'CONCATENATE\s*\(',
            'SUM': r'SUM\s*\(',
            'AVERAGE': r'AVERAGE\s*\(',
            'MAX': r'MAX\s*\(',
            'MIN': r'MIN\s*\(',
            'ROUND': r'ROUND\s*\(',
            'TODAY': r'TODAY\s*\(',
            'NOW': r'NOW\s*\(',
            'DATE': r'DATE\s*\(',
            'INDIRECT': r'INDIRECT\s*\(',
            'OFFSET': r'OFFSET\s*\(',
            'CHOOSE': r'CHOOSE\s*\(',
            'SWITCH': r'SWITCH\s*\(',
            'TEXTJOIN': r'TEXTJOIN\s*\(',
            'FILTER': r'FILTER\s*\(',
            'UNIQUE': r'UNIQUE\s*\(',
            'SORT': r'SORT\s*\(',
            'CUSTOM': r'[A-Z_][A-Z0-9_]*\s*\('
        }
    
    def extract_formulas(self, workbook: openpyxl.Workbook) -> List[Dict[str, Any]]:
        """
        Extract formulas from specific target sheets only
        
        Args:
            workbook: openpyxl Workbook object
            
        Returns:
            List of dictionaries containing formula information
        """
        formulas = []
        target_sheets = ['product-pre-release', 'pre-release-version']
        
        try:
            for sheet_name in workbook.sheetnames:
                if sheet_name in target_sheets:
                    sheet = workbook[sheet_name]
                    sheet_formulas = self._extract_sheet_formulas(sheet, sheet_name)
                    formulas.extend(sheet_formulas)
                
        except Exception as e:
            raise Exception(f"Error extracting formulas: {str(e)}")
        
        return formulas
    
    def extract_b5_cell(self, workbook: openpyxl.Workbook) -> Dict[str, Any]:
        """
        Extract specifically the B5 cell from pre-release-version sheet
        
        Args:
            workbook: openpyxl Workbook object
            
        Returns:
            Dictionary containing B5 cell information
        """
        try:
            if 'pre-release-version' in workbook.sheetnames:
                sheet = workbook['pre-release-version']
                cell = sheet['B5']
                
                if cell.data_type == 'f' and cell.value:  # Formula cell
                    return self._analyze_formula(cell.value, 'B5', 'pre-release-version')
                else:
                    return {
                        'sheet': 'pre-release-version',
                        'cell': 'B5',
                        'formula': None,
                        'value': cell.value,
                        'description': 'No formula found in B5 cell'
                    }
            else:
                return {
                    'error': 'Sheet "pre-release-version" not found in workbook'
                }
        except Exception as e:
            return {
                'error': f'Error extracting B5 cell: {str(e)}'
            }
    
    def extract_service_names(self, workbook: openpyxl.Workbook) -> List[Dict[str, Any]]:
        """
        Extract service names from both target sheets by looking at cell values
        
        Args:
            workbook: openpyxl Workbook object
            
        Returns:
            List of service information dictionaries
        """
        services = []
        target_sheets = ['product-pre-release', 'pre-release-version']
        
        try:
            for sheet_name in target_sheets:
                if sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    
                    # Look through all cells for any content
                    for row in sheet.iter_rows(min_row=1, max_row=100):
                        for cell in row:
                            if cell.value:
                                cell_value = str(cell.value).strip()
                                # Look for service-like names
                                if self._is_service_name(cell_value):
                                    services.append({
                                        'service_name': cell_value,
                                        'sheet': sheet_name,
                                        'cell': f"{cell.column_letter}{cell.row}",
                                        'source': 'cell_value'
                                    })
                                    
        except Exception as e:
            print(f"Warning: Error extracting service names: {str(e)}")
        
        # Remove duplicates
        unique_services = []
        seen_names = set()
        for service in services:
            if service['service_name'] not in seen_names:
                seen_names.add(service['service_name'])
                unique_services.append(service)
        
        return unique_services
    
    def get_all_cell_data(self, workbook: openpyxl.Workbook) -> Dict[str, List[Dict[str, Any]]]:
        """
        Extract all cell data from target sheets for debugging
        """
        all_data = {}
        target_sheets = ['product-pre-release', 'pre-release-version']
        
        for sheet_name in target_sheets:
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_data = []
                
                for row in sheet.iter_rows(min_row=1, max_row=20, min_col=1, max_col=10):
                    for cell in row:
                        if cell.value:
                            sheet_data.append({
                                'cell': f"{cell.column_letter}{cell.row}",
                                'value': str(cell.value),
                                'data_type': cell.data_type
                            })
                
                all_data[sheet_name] = sheet_data
        
        return all_data
    
    def _is_service_name(self, text: str) -> bool:
        """
        Check if text looks like a service name
        """
        if len(text) < 3:
            return False
            
        # Common service name patterns
        service_patterns = [
            'studio-backend', 'studio-ui', 'bodhee-core', 'file-upload-connector',
            'bodhee-security', 'bxs-masterdata', 'bxs-masterdata-management'
        ]
        
        # Check if it matches known patterns
        for pattern in service_patterns:
            if pattern in text.lower():
                return True
                
        # Check if it has service-like structure (word-word or word_word)
        if '-' in text or '_' in text:
            parts = text.replace('_', '-').split('-')
            if len(parts) >= 2 and all(len(part) > 1 for part in parts):
                return True
                
        return False
    
    def _extract_sheet_formulas(self, sheet, sheet_name: str) -> List[Dict[str, Any]]:
        """
        Extract formulas from a specific sheet
        
        Args:
            sheet: openpyxl Worksheet object
            sheet_name: Name of the sheet
            
        Returns:
            List of formula dictionaries
        """
        formulas = []
        
        try:
            # Iterate through all cells in the sheet
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.data_type == 'f' and cell.value:  # Formula cell
                        formula_info = self._analyze_formula(
                            cell.value, 
                            f"{cell.column_letter}{cell.row}", 
                            sheet_name
                        )
                        if formula_info:
                            formulas.append(formula_info)
                            
        except Exception as e:
            print(f"Warning: Error processing sheet {sheet_name}: {str(e)}")
        
        return formulas
    
    def _analyze_formula(self, formula, cell_address: str, sheet_name: str) -> Dict[str, Any]:
        """
        Analyze a formula and extract information
        
        Args:
            formula: The formula (string or ArrayFormula object)
            cell_address: Cell address (e.g., A1)
            sheet_name: Name of the sheet
            
        Returns:
            Dictionary with formula information
        """
        try:
            # Handle different formula types
            if hasattr(formula, 'text'):
                # ArrayFormula object
                formula_str = str(formula.text) if formula.text else str(formula)
            else:
                # Regular string formula
                formula_str = str(formula)
            
            # Clean the formula (remove leading =)
            clean_formula = formula_str.lstrip('=')
            
            # Determine formula type
            formula_type = self._identify_formula_type(clean_formula)
            
            # Extract references
            references = self._extract_references(clean_formula)
            
            # Generate description
            description = self._generate_description(clean_formula, formula_type)
            
            return {
                'sheet': sheet_name,
                'cell': cell_address,
                'formula': formula_str,
                'clean_formula': clean_formula,
                'formula_type': formula_type,
                'references': references,
                'description': description,
                'complexity': self._assess_complexity(clean_formula)
            }
            
        except Exception as e:
            formula_str = str(formula) if formula else "Empty formula"
            return {
                'sheet': sheet_name,
                'cell': cell_address,
                'formula': formula_str,
                'clean_formula': formula_str.lstrip('=') if formula_str else "",
                'formula_type': 'UNKNOWN',
                'references': [],
                'description': f"Error analyzing formula: {str(e)}",
                'complexity': 'unknown'
            }
    
    def _identify_formula_type(self, formula: str) -> str:
        """
        Identify the primary type of formula
        
        Args:
            formula: Clean formula string
            
        Returns:
            Formula type string
        """
        formula_upper = formula.upper()
        
        # Check for specific function patterns
        for func_type, pattern in self.formula_patterns.items():
            if re.search(pattern, formula_upper):
                if func_type == 'CUSTOM':
                    # Try to extract the actual function name
                    match = re.search(r'([A-Z_][A-Z0-9_]*)\s*\(', formula_upper)
                    if match:
                        return match.group(1)
                return func_type
        
        # Check for basic arithmetic
        if any(op in formula for op in ['+', '-', '*', '/', '^']):
            return 'ARITHMETIC'
        
        # Check for simple cell references
        if re.match(r'^[A-Z]+[0-9]+$', formula.replace('$', '')):
            return 'REFERENCE'
        
        return 'UNKNOWN'
    
    def _extract_references(self, formula: str) -> List[str]:
        """
        Extract cell and range references from formula
        
        Args:
            formula: Clean formula string
            
        Returns:
            List of references found
        """
        references = []
        
        # Pattern for cell references (A1, $A$1, Sheet1!A1, etc.)
        cell_pattern = r'(?:[A-Za-z_][A-Za-z0-9_]*!)?[$]?[A-Z]+[$]?[0-9]+'
        
        # Pattern for range references (A1:B2, Sheet1!A1:B2, etc.)
        range_pattern = r'(?:[A-Za-z_][A-Za-z0-9_]*!)?[$]?[A-Z]+[$]?[0-9]+:[$]?[A-Z]+[$]?[0-9]+'
        
        # Find all references
        cell_refs = re.findall(cell_pattern, formula, re.IGNORECASE)
        range_refs = re.findall(range_pattern, formula, re.IGNORECASE)
        
        references.extend(cell_refs)
        references.extend(range_refs)
        
        # Remove duplicates while preserving order
        seen = set()
        unique_refs = []
        for ref in references:
            if ref not in seen:
                seen.add(ref)
                unique_refs.append(ref)
        
        return unique_refs
    
    def _generate_description(self, formula: str, formula_type: str) -> str:
        """
        Generate a human-readable description of the formula
        
        Args:
            formula: Clean formula string
            formula_type: Type of formula
            
        Returns:
            Description string
        """
        descriptions = {
            'VLOOKUP': 'Vertical lookup function to find values in a table',
            'HLOOKUP': 'Horizontal lookup function to find values in a table',
            'INDEX': 'Returns a value from a specific position in a range',
            'MATCH': 'Finds the position of a value in a range',
            'IF': 'Conditional logic function',
            'SUMIF': 'Conditional sum function',
            'COUNTIF': 'Conditional count function',
            'CONCATENATE': 'Text concatenation function',
            'SUM': 'Summation function',
            'AVERAGE': 'Average calculation function',
            'MAX': 'Maximum value function',
            'MIN': 'Minimum value function',
            'ROUND': 'Number rounding function',
            'TODAY': 'Current date function',
            'NOW': 'Current date and time function',
            'DATE': 'Date construction function',
            'INDIRECT': 'Indirect reference function',
            'OFFSET': 'Dynamic reference function',
            'CHOOSE': 'Value selection function',
            'SWITCH': 'Multi-condition selection function',
            'TEXTJOIN': 'Text joining function with delimiter',
            'FILTER': 'Dynamic array filtering function',
            'UNIQUE': 'Returns unique values from a range',
            'SORT': 'Sorts data in a range',
            'ARITHMETIC': 'Mathematical calculation',
            'REFERENCE': 'Simple cell reference'
        }
        
        base_description = descriptions.get(formula_type, 'Custom or complex formula')
        
        # Add reference count if available
        ref_count = len(self._extract_references(formula))
        if ref_count > 0:
            base_description += f" (references {ref_count} cell{'s' if ref_count != 1 else ''})"
        
        return base_description
    
    def _assess_complexity(self, formula: str) -> str:
        """
        Assess the complexity of a formula
        
        Args:
            formula: Clean formula string
            
        Returns:
            Complexity level string
        """
        # Count various complexity indicators
        nested_functions = len(re.findall(r'\w+\s*\(', formula))
        references = len(self._extract_references(formula))
        length = len(formula)
        
        # Calculate complexity score
        complexity_score = 0
        complexity_score += nested_functions * 2
        complexity_score += references
        complexity_score += length // 50
        
        if complexity_score <= 3:
            return 'low'
        elif complexity_score <= 8:
            return 'medium'
        else:
            return 'high'
    
    def get_formula_statistics(self, formulas: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        Generate statistics about the extracted formulas
        
        Args:
            formulas: List of formula dictionaries
            
        Returns:
            Statistics dictionary
        """
        if not formulas:
            return {}
        
        stats = {
            'total_formulas': len(formulas),
            'sheets': len(set(f['sheet'] for f in formulas)),
            'formula_types': {},
            'complexity_distribution': {'low': 0, 'medium': 0, 'high': 0, 'unknown': 0},
            'average_references': 0,
            'most_complex_formula': None
        }
        
        # Count formula types
        for formula in formulas:
            formula_type = formula['formula_type']
            stats['formula_types'][formula_type] = stats['formula_types'].get(formula_type, 0) + 1
            
            # Count complexity
            complexity = formula['complexity']
            stats['complexity_distribution'][complexity] += 1
        
        # Calculate average references
        total_refs = sum(len(f['references']) for f in formulas)
        stats['average_references'] = total_refs / len(formulas) if formulas else 0
        
        # Find most complex formula
        high_complexity = [f for f in formulas if f['complexity'] == 'high']
        if high_complexity:
            stats['most_complex_formula'] = max(high_complexity, key=lambda x: len(x['formula']))
        
        return stats
