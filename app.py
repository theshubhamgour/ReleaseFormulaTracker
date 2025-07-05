import streamlit as st
import openpyxl
import io
import time
from formula_processor import FormulaProcessor
from stack_generator import StackGenerator

st.set_page_config(page_title="Release Configuration Tracker", layout="wide")

def main():
    st.title("Release Configuration Tracker")
    st.write("Upload an Excel workbook to extract formulas and generate release image stacks")
    
    # Initialize session state
    if 'processed_file' not in st.session_state:
        st.session_state.processed_file = None
    if 'formulas_data' not in st.session_state:
        st.session_state.formulas_data = None
    if 'b5_cell_data' not in st.session_state:
        st.session_state.b5_cell_data = None
    if 'service_names' not in st.session_state:
        st.session_state.service_names = None
    if 'release_versions' not in st.session_state:
        st.session_state.release_versions = []
    if 'workbook' not in st.session_state:
        st.session_state.workbook = None
    
    # File upload
    uploaded_file = st.file_uploader("Choose an Excel file", type=['xlsx'])
    
    if uploaded_file is not None:
        if st.session_state.processed_file != uploaded_file.name:
            # Show progress bar
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            try:
                # Step 1: Load workbook
                status_text.text("Loading Excel file...")
                progress_bar.progress(20)
                time.sleep(0.3)
                
                workbook_bytes = io.BytesIO(uploaded_file.read())
                workbook = openpyxl.load_workbook(workbook_bytes, data_only=False)
                st.session_state.workbook = workbook
                
                # Step 2: Extract release versions
                status_text.text("Extracting release versions...")
                progress_bar.progress(40)
                time.sleep(0.3)
                
                release_versions = extract_release_versions(workbook)
                st.session_state.release_versions = release_versions
                
                # Step 3: Process formulas
                status_text.text("Processing formulas...")
                progress_bar.progress(60)
                time.sleep(0.3)
                
                processor = FormulaProcessor()
                formulas_data = processor.extract_formulas(workbook)
                b5_cell_data = processor.extract_b5_cell(workbook)
                service_names = processor.extract_service_names(workbook)
                
                # Step 4: Store results
                status_text.text("Finalizing processing...")
                progress_bar.progress(80)
                time.sleep(0.3)
                
                st.session_state.formulas_data = formulas_data
                st.session_state.b5_cell_data = b5_cell_data
                st.session_state.service_names = service_names
                st.session_state.processed_file = uploaded_file.name
                
                # Complete
                progress_bar.progress(100)
                status_text.text("Processing complete!")
                time.sleep(0.5)
                
                # Clear progress indicators
                progress_bar.empty()
                status_text.empty()
                
                st.success("Excel file processed successfully!")
                
                # Show available sheets for debugging
                if workbook:
                    st.info(f"Available sheets: {', '.join(workbook.sheetnames)}")
                
                if release_versions:
                    st.info(f"Found {len(release_versions)} release versions")
                else:
                    st.warning("No release versions found in 'product-pre-release-neewee' sheet starting from B6")
                
            except Exception as e:
                progress_bar.empty()
                status_text.empty()
                st.error(f"Error processing file: {str(e)}")
    
    # Release version selection and stack generation
    if st.session_state.processed_file and st.session_state.release_versions:
        st.subheader("Release Version Selection")
        
        # Dropdown for release version selection
        selected_version = st.selectbox(
            "Select Release Version",
            st.session_state.release_versions,
            key="version_selector"
        )
        
        # Generate stack button
        if st.button("Generate Stack", type="primary"):
            if selected_version:
                with st.spinner("Generating stack..."):
                    try:
                        # Update B5 cell with selected version
                        updated_workbook = update_b5_cell(st.session_state.workbook, selected_version)
                        
                        # Process the updated workbook to get stack data
                        processor = FormulaProcessor()
                        stack_data = extract_stack_data(updated_workbook)
                        
                        st.success("Stack generated successfully!")
                        
                        # Display results in table format
                        if stack_data:
                            st.subheader("Generated Stack")
                            
                            # Create table with Service Name and Docker Image
                            st.table(stack_data)
                            
                        else:
                            st.warning("No services found in the generated stack")
                            
                    except Exception as e:
                        st.error(f"Error generating stack: {str(e)}")
            else:
                st.warning("Please select a release version first")

def extract_release_versions(workbook):
    """Extract release versions from 'product-pre-release-neewee' sheet starting from B6"""
    try:
        sheet_name = 'product-pre-release-neewee'
        if sheet_name not in workbook.sheetnames:
            return []
        
        sheet = workbook[sheet_name]
        versions = []
        
        # Start from B6 and go down until we hit an empty cell
        row = 6
        while True:
            cell_value = sheet[f'B{row}'].value
            if cell_value is None or str(cell_value).strip() == '':
                break
            versions.append(str(cell_value).strip())
            row += 1
            
            # Safety limit to prevent infinite loops
            if row > 1000:
                break
        
        return versions
    except Exception as e:
        st.error(f"Error extracting release versions: {str(e)}")
        return []

def update_b5_cell(workbook, selected_version):
    """Update B5 cell in 'pre-release-version' sheet with selected version"""
    try:
        sheet_name = 'pre-release-version'
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet['B5'] = selected_version
            # Force workbook to recalculate
            workbook.active = sheet
        return workbook
    except Exception as e:
        st.error(f"Error updating B5 cell: {str(e)}")
        return workbook

def extract_stack_data(workbook):
    """Extract stack data from the workbook after B5 is updated"""
    try:
        sheet_name = 'pre-release-version'
        if sheet_name not in workbook.sheetnames:
            return []
        
        sheet = workbook[sheet_name]
        stack_data = []
        
        # Look for service names and corresponding data in the sheet
        # This is a simplified approach - you may need to adjust based on your actual Excel structure
        for row in range(1, 100):  # Check first 100 rows
            for col in range(1, 10):  # Check first 10 columns
                cell_value = sheet.cell(row=row, column=col).value
                if cell_value and isinstance(cell_value, str):
                    cell_value = str(cell_value).strip()
                    # Look for service-like names (containing hyphens or specific patterns)
                    if is_service_name(cell_value):
                        # Generate Docker image name
                        docker_image = generate_docker_image(cell_value, sheet['B5'].value)
                        stack_data.append({
                            'Service Name': cell_value,
                            'Docker Image': docker_image
                        })
        
        # Remove duplicates
        seen_services = set()
        unique_stack_data = []
        for item in stack_data:
            if item['Service Name'] not in seen_services:
                seen_services.add(item['Service Name'])
                unique_stack_data.append(item)
        
        return unique_stack_data
    except Exception as e:
        st.error(f"Error extracting stack data: {str(e)}")
        return []

def is_service_name(text):
    """Check if text looks like a service name"""
    if not text or len(text) < 3:
        return False
    
    # Look for service-like patterns
    service_patterns = [
        '-service', '-api', '-worker', '-processor', '-handler',
        'service-', 'api-', 'worker-', 'processor-', 'handler-'
    ]
    
    text_lower = text.lower()
    return any(pattern in text_lower for pattern in service_patterns) or '-' in text

def generate_docker_image(service_name, version):
    """Generate Docker image name based on service and version"""
    if not version:
        version = "latest"
    
    # Clean version if needed
    version_clean = str(version).replace('v', '').replace('-pre', '').strip()
    if version_clean.endswith('.'):
        version_clean = version_clean[:-1]
    
    # Clean service name
    service_clean = service_name.replace('_', '').lower()
    
    return f"neewee/{service_clean}:pre-release-v{version_clean}"

if __name__ == "__main__":
    main()